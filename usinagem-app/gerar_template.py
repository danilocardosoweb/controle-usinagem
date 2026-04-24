#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os

# Tentar importar openpyxl, se não existir, criar um arquivo CSV alternativo
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Criar workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove a aba padrão
    
    # ===== ABA 1: INSTRUÇÕES =====
    ws_instr = wb.create_sheet('Instruções', 0)
    
    instructions = [
        ['TEMPLATE DE APONTAMENTOS DE USINAGEM'],
        [''],
        ['INSTRUÇÕES DE PREENCHIMENTO:'],
        [''],
        ['1. CAMPOS OBRIGATÓRIOS (devem ser preenchidos):'],
        ['   - Data Início: formato DD/MM/YYYY HH:MM (ex: 18/03/2026 14:55)'],
        ['   - Data Fim: formato DD/MM/YYYY HH:MM (ex: 18/03/2026 15:55)'],
        ['   - Pedido/Seq: código do pedido (ex: 84659/10)'],
        ['   - Máquina: nome da máquina (consulte a aba "Referência")'],
        ['   - Quantidade: número de peças produzidas'],
        [''],
        ['2. CAMPOS OPCIONAIS:'],
        ['   - Qtd Refugo: quantidade de peças refugadas'],
        ['   - Comprimento Refugo: comprimento em mm'],
        ['   - Dureza: valor de dureza do material'],
        ['   - Rack/Pallet: identificação do rack ou pallet'],
        ['   - Observações: anotações adicionais'],
        [''],
        ['3. CAMPOS PREENCHIDOS AUTOMATICAMENTE:'],
        ['   - Operador: será preenchido com o usuário logado'],
        ['   - Produto: extraído do pedido'],
        ['   - Cliente: extraído do pedido'],
        ['   - Perfil Longo: extraído do pedido'],
        [''],
        ['4. DATAS E HORAS:'],
        ['   - Use formato DD/MM/YYYY HH:MM'],
        ['   - A hora de fim NÃO pode ser anterior à hora de início'],
        ['   - Exemplo correto: 18/03/2026 14:55 até 18/03/2026 15:55'],
        [''],
        ['5. MÁQUINAS DISPONÍVEIS:'],
        ['   - Consulte a aba "Referência" para ver a lista de máquinas'],
        [''],
        ['6. APÓS PREENCHER:'],
        ['   - Salve o arquivo em formato Excel (.xlsx)'],
        ['   - Importe no app em "Apontamentos > Importar Planilha"'],
        ['   - O app validará e carregará os dados automaticamente'],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instr.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
    
    ws_instr.column_dimensions['A'].width = 80
    
    # ===== ABA 2: REFERÊNCIA =====
    ws_ref = wb.create_sheet('Referência', 1)
    
    reference_data = [
        ['MÁQUINAS DISPONÍVEIS', '', 'OPERADORES'],
        ['Nome da Máquina', 'Código', 'Nome do Operador'],
        ['Serra Doppia 2 cabeças', 'SERRA_DOPPIA_2', 'Danilo Cardoso'],
        ['Torno CNC', 'TORNO_CNC', 'João Silva'],
        ['Fresadora', 'FRESADORA', 'Maria Santos'],
        ['Retífica', 'RETIFICA', 'Pedro Oliveira'],
        ['', '', ''],
        ['DICAS:', '', ''],
        ['- Use o nome exato da máquina ao preencher', '', ''],
        ['- Se a máquina não está na lista, contate o administrador', '', ''],
    ]
    
    for row_idx, row_data in enumerate(reference_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_ref.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=12)
            elif row_idx == 2:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    ws_ref.column_dimensions['A'].width = 25
    ws_ref.column_dimensions['B'].width = 20
    ws_ref.column_dimensions['C'].width = 25
    
    # ===== ABA 3: ENTRADA DE DADOS =====
    ws_data = wb.create_sheet('Apontamentos', 2)
    
    headers = [
        'Data Início',
        'Data Fim',
        'Pedido/Seq',
        'Máquina',
        'Quantidade',
        'Qtd Refugo',
        'Comprimento Refugo (mm)',
        'Dureza',
        'Rack/Pallet',
        'Observações'
    ]
    
    # Adicionar cabeçalho
    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Adicionar linha de exemplo
    example_data = [
        '18/03/2026 14:55',
        '18/03/2026 15:55',
        '84659/10',
        'Serra Doppia 2 cabeças',
        '8000',
        '50',
        '1340',
        '58',
        '4117',
        'Troca de pallet 10:20/10:40'
    ]
    
    for col_idx, value in enumerate(example_data, 1):
        cell = ws_data.cell(row=2, column=col_idx, value=value)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.font = Font(italic=True, color='595959')
    
    # Adicionar 20 linhas vazias para preenchimento
    for row_idx in range(3, 23):
        for col_idx in range(1, 11):
            ws_data.cell(row=row_idx, column=col_idx, value='')
    
    # Definir largura das colunas
    widths = [18, 18, 15, 25, 12, 12, 20, 10, 15, 30]
    for col_idx, width in enumerate(widths, 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Salvar arquivo
    output_path = r'c:\Users\Danilo\Desktop\apps\Usinagem\TEMPLATE_APONTAMENTOS.xlsx'
    wb.save(output_path)
    
    print(f'✅ Template Excel criado com sucesso!')
    print(f'📁 Arquivo: {output_path}')
    sys.exit(0)

except ImportError:
    print('❌ openpyxl não instalado. Criando arquivo CSV alternativo...')
    
    # Criar arquivo CSV como alternativa
    csv_path = r'c:\Users\Danilo\Desktop\apps\Usinagem\TEMPLATE_APONTAMENTOS.csv'
    
    csv_content = '''Data Início,Data Fim,Pedido/Seq,Máquina,Quantidade,Qtd Refugo,Comprimento Refugo (mm),Dureza,Rack/Pallet,Observações
18/03/2026 14:55,18/03/2026 15:55,84659/10,Serra Doppia 2 cabeças,8000,50,1340,58,4117,Troca de pallet 10:20/10:40
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
,,,,,,,,
'''
    
    with open(csv_path, 'w', encoding='utf-8') as f:
        f.write(csv_content)
    
    print(f'✅ Template CSV criado com sucesso!')
    print(f'📁 Arquivo: {csv_path}')
    print('⚠️  Para usar como Excel, abra o arquivo CSV no Excel e salve como .xlsx')
    sys.exit(0)
